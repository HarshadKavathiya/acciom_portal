import ast
import json
from io import BytesIO

from flasgger import swag_from
from flask import current_app as app
from flask import request, Response
from flask_jwt_extended import jwt_required, get_jwt_identity
from flask_restful import Resource, reqparse
from openpyxl import load_workbook

from application.api.dbdetails import create_dbconnection
from application.common.Response import success, error
# from celery_task import my_background_task
from application.helper.runner_class import run_by_case_id
from application.helper.runner_class import split_db
from application.models.user import TestSuite, TestCase, TestCaseLog, DbDetail

parser = reqparse.RequestParser()
parser.add_argument('sheet',
                    help='This field cannot be blank',
                    required=True)
parser.add_argument('selectedcase',
                    help='this field is required',
                    required=True)
parser.add_argument('suitename',
                    help='this field is required',
                    required=True)
parser.add_argument('exvalue',
                    help='this field is required',
                    required=True)


def args_as_list(s):
    v = ast.literal_eval(s)
    if type(v) is not list:
        pass
    return v


class TestSuites(Resource):
    @jwt_required
    @swag_from('/application/apidocs/test_suitepost.yml')
    def post(self):
        current_user = get_jwt_identity()
        data = parser.parse_args()
        sheet = data['sheet']
        file = request.files['inputFile']
        suite_name = data['suitename']
        user_id = get_jwt_identity()
        temp_file = TestSuite(user_id=user_id,
                              excel_name=file.filename,
                              test_suite_name=suite_name)

        temp_file.save_to_db()
        wb = load_workbook(filename=BytesIO(file.read()))
        sheet_index = wb.sheetnames.index(sheet)
        ws = wb.worksheets[sheet_index]
        temp_test1 = [str(i - 2) for i in range(2, ws.max_row + 1)]
        # row+1 to avoid index out of range error
        print(temp_test1)
        temp_test = []
        for i in range(0, ws.max_column):
            if (str(ws[1][i].value) != 'None'):
                temp_test.append([str(ws[x][i].value)
                                  for x in range(2, ws.max_row)])
        data = parser.parse_args()
        print(data['selectedcase'])
        test_case_list = str(data['selectedcase']).split(",")
        print(len(test_case_list))
        i = 0
        for j in range(ws.max_row - 1):
            print("73", temp_test1[j])
            print('74', test_case_list)
            print(temp_test1[j] in test_case_list)
            if temp_test1[j] in test_case_list:
                test_case_list.remove(temp_test1[j])
                db_list = split_db(temp_test[i + 2][j])
                src_db_id = create_dbconnection(current_user,
                                                db_list[2].lower(),
                                                db_list[0],
                                                db_list[4].lower(),
                                                db_list[6])
                target_db_id = create_dbconnection(current_user,
                                                   db_list[3].lower(),
                                                   db_list[1],
                                                   db_list[5].lower(),
                                                   db_list[7])
                jsondict = {}
                columndata = temp_test[i + 4][j]
                column = {}
                if columndata == "None" or columndata.isspace():
                    pass
                else:
                    if ":" in columndata:
                        x = temp_test[i + 4][j].split(":")
                        z = []
                        for m in x:
                            y = m.split(";")
                            z.append(y)
                        column = {}
                        for n in range(len(z[0])):
                            column[z[0][n]] = z[1][n]

                    else:
                        columnlist = temp_test[i + 4][j].split(";")
                        column = {}
                        for k in range(len(columnlist)):
                            column[columnlist[k]] = columnlist[k]

                tablelist = temp_test[i + 3][j].split(":")
                table = {}
                table[tablelist[0]] = tablelist[1]
                query = {}
                p = temp_test[i + 5][j]
                if p == "None" or p.isspace():
                    pass
                else:
                    if ";" in p:
                        query_split = p.split(";")
                        final = [a.split(":") for a in query_split]
                        # Added logic to avoid index error if only srcqry is given
                        query["sourceqry"] = final[0][1] if 'srcqry' in final[0] else ""
                        query["targetqry"] = final[1][1] if 'targetqry' in final[1] else ""
                    else:
                        q = p.strip("targetqry:")
                        query["targetqry"] = q

                jsondict = {"column": column, "table": table, "query": query}

                temp = TestCase(test_suite_id=temp_file.test_suite_id,
                                test_id=temp_test[i + 1][j],
                                test_status=0,
                                test_case_detail=json.dumps(jsondict),
                                test_name=temp_test[i][j],
                                src_db_id=src_db_id,
                                target_db_id=target_db_id)
                temp.save_to_db()
        if int(data['exvalue']) == 1:
            test_suite = TestSuite.query.filter_by(
                test_suite_id=temp.test_suite_id).first()
            for each_test in test_suite.test_case:
                run_by_case_id(each_test.test_case_id)
        return {'message': 'data saved to database'}

    @jwt_required
    @swag_from('/application/apidocs/test_suiteget.yml')
    def get(self):
        try:
            uid = get_jwt_identity()
            return {"suites": TestSuite.return_all(uid),
                    "success": True}
        except Exception as e:
            app.logger.debug(str(e))
            return {"success": False, "message": str(e)}


class TestCaseLogDetail(Resource):
    @jwt_required
    @swag_from('/application/apidocs/testcaselogdetail.yml')
    def get(self, test_case_log_id):
        return {"test_case_log": TestCaseLog.return_all_log(test_case_log_id),
                "success": True}


class ExportTestLog(Resource):
    @swag_from('/application/apidocs/exporttestlog.yml')
    def get(self, case_log_id):
        case_log = TestCaseLog.query.filter_by(
            test_case_log_id=case_log_id).first()
        test_case = case_log.test_cases
        response = []
        if test_case.test_name == 'Datavalidation':
            data = ast.literal_eval(case_log.src_execution_log)

            dict_key = ast.literal_eval(data[0])
            key_list = [key for key in dict_key.keys()]
            response.append(key_list)
            for i in range(len(data)):
                value_list = [x for x in ast.literal_eval(data[i]).values()]
                response.append(value_list)
            response = json.dumps(response)
        elif test_case.test_name == 'CountCheck':
            src_response = case_log.src_execution_log
            des_response = case_log.des_execution_log
            res = [['Source Count', 'destination Count']]
            res.append([src_response, des_response])
            response = json.dumps(res)
        elif test_case.test_name == 'DuplicateCheck' or 'NullCheck':
            response = case_log.des_execution_log

        from openpyxl import Workbook
        from tempfile import NamedTemporaryFile
        wb = Workbook()
        ws = wb.active

        response = json.loads(response)
        for each in response:
            ws.append(list(each))

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        return Response(
            stream,
            mimetype="application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet",
            headers={"Content-disposition": "attachment; "
                                            "filename=export123.xlsx"})


class ConnectionDetails(Resource):
    '''
    accepts suite_id of user and returns
     corresponding case_ids and connection_ids of user
    '''

    @jwt_required
    @swag_from('/application/apidocs/connectiondetails.yml')
    def get(self, suite_id):
        try:
            current_user = get_jwt_identity()
            db_obj = DbDetail.query.filter_by(user_id=current_user).all()
            suite_obj = TestSuite. \
                query.filter_by(test_suite_id=suite_id).first()
            all_case = [(i.test_case_id, i.test_name)
                        for i in suite_obj.test_case]
            all_connection = [(i.db_id, i.connection_name) for i in db_obj]
            return {"all_connections": all_connection, "all_cases": all_case}
        except Exception as e:
            app.logger.error(str(e))
            return {"success": False, "message": str(e)}


class SelectConnection(Resource):
    @jwt_required
    @swag_from('/application/apidocs/selectconnection.yml')
    def post(self):
        try:
            parser = reqparse.RequestParser()
            parser.add_argument('connection_type',
                                help='This field cannot be blank',
                                required=True)
            parser.add_argument('case_id',
                                help='This field cannot be blank',
                                required=True,
                                type=args_as_list, default=[])
            parser.add_argument('db_id',
                                help='This field cannot be blank',
                                required=True)
            data = parser.parse_args()

            if data['connection_type'] == "source":
                for i in data['case_id']:
                    testcase = TestCase.query.filter_by(test_case_id=i).first()
                    testcase.src_db_id = data["db_id"]
                    testcase.save_to_db()

            elif data['connection_type'] == "dest":
                for i in data['case_id']:
                    testcase = TestCase.query.filter_by(test_case_id=i).first()
                    testcase.target_db_id = data["db_id"]
                    testcase.save_to_db()
            return success({"success": True, "message": "Updated Details"})
        except Exception as e:
            app.loggger.error(e)
            return error({"success": False, "message": str(e)})
