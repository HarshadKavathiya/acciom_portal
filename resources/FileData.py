from flask_restful import Resource, reqparse
from flask import request
from models.user import TestSuite, TestCase
from flask_jwt_extended import jwt_required, get_jwt_identity
from openpyxl import load_workbook
from io import BytesIO
from celery_task import my_background_task
from test.runner_class import run_by_case_id

parser = reqparse.RequestParser()
parser.add_argument('sheet', help='This field cannot be blank', required=True)
parser.add_argument('selectedcase', help='this field is required', required=True)
parser.add_argument('suitename', help='this field is required', required=True)
parser.add_argument('exvalue', help='this field is required', required=True)


class Upload(Resource):
    @jwt_required
    def post(self):
        data = parser.parse_args()
        sheet = data['sheet']
        selected_case = data['selectedcase']
        file = request.files['inputFile']
        suite_name = data['suitename']
        print(suite_name)
        print("execution value", data['exvalue'])
        user_id = get_jwt_identity()
        print(user_id)
        temp_file = TestSuite(user_id=user_id, excel_name=file.filename,test_suite_name=suite_name)

        temp_file.save_to_db()
        print(temp_file.test_suite_id)

        print(selected_case,  sheet)
        wb = load_workbook(filename=BytesIO(file.read()))
        sheet_index = wb.sheetnames.index(sheet)
        ws = wb.worksheets[sheet_index]
        # print(ws)
        temp_test = [[str(ws[x][0].value) for x in range(2, ws.max_row+1)]] #from column 2nd
        for i in range(1, ws.max_column):
            temp_test.append([str(ws[x][i].value) for x in range(2, ws.max_row+1)])
            data = parser.parse_args()

        test_case_list = str(data['selectedcase']).split(",")
        # print(test_case_list)

        i = 0
        # # TODO: remove the hardcode colummn numbers and retrive the indexes of column
        # print(temp_test)
        for j in range(ws.max_row-1):
            if temp_test[i][j] in test_case_list:
                temp = TestCase(test_suite_id=temp_file.test_suite_id
                                , test_id=temp_test[i][j], test_status=0
                                , test_priority=temp_test[i+2][j]
                                , test_detail=temp_test[i+3][j]
                                , test_column=temp_test[i+4][j]
                                , table_src_target=temp_test[i+5][j]
                                , test_name=temp_test[i+6][j]
                                , test_queries=temp_test[i+7][j]
                                , test_expected=temp_test[i+8][j]
                                , test_actual=temp_test[i+9][j]
                                , test_created_by=temp_test[i+10][j]
                                , test_executed_by=temp_test[i+11][j]
                                , test_comment=temp_test[i+12][j])
                temp.save_to_db()

        if int(data['exvalue']) == 1:
            # my_background_task.apply_async(countdown=15)
            print(temp.test_suite_id)
            test_suite = TestSuite.query.filter_by(test_suite_id=temp.test_suite_id).first()
            for each_test in test_suite.test_case:
                # TODO:my_background_task.apply_async(args=[each_test.test_case_id])
                run_by_case_id(each_test.test_case_id)
        return {'message': 'data saved to database'}


class GetUpload(Resource):
    def get(self, user_id):
        arr1 = []
        # print(TestSuite.return_all(user_id))

        return {"suites": TestSuite.return_all(user_id),
                "success": True}






